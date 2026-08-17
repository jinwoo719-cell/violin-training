"""
기능 목록 → 엑셀.

features.py 가 모은 줄들을 시트 네 장으로 씁니다.

    개요        숫자 요약 (전부 수식 — 목록이 바뀌면 같이 바뀝니다)
    기능목록    기능 한 줄씩 · 모듈명 포함
    모듈        모듈이 무엇을 맡는가 + 기능 수 (수식)
    다음단계    아직 안 만든 것

숫자는 손으로 적지 않고 **수식**으로 둡니다.
기능목록에 한 줄만 더해도 개요와 모듈 시트가 저절로 맞습니다.
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

#| 흐름  기능 목록 → 시트 네 장 (개요 · 기능목록 · 모듈 · 다음단계)

FONT = "맑은 고딕"                  # 윈도우 한글 기본. 한글이 깨지지 않습니다
INK = "1F2430"
HEAD_BG = "1F2430"
BAND = "F4F6FA"
LINE = "C9CFDA"

_thin = Side(style="thin", color=LINE)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _head(ws, cols, row=1):
    """머리글 한 줄."""
    #| 흐름  머리글을 쓰고 색과 테두리를 입힌다
    for c, name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 26


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_features(wb, rows, cols):
    """기능목록 — 이 파일의 본체."""
    #| 흐름  기능 한 줄씩 쓰고, 모듈이 바뀔 때마다 옅은 띠를 넣는다
    #| 입력  기능 목록 · 열 이름
    #| 단계  머리글을 쓰고 틀을 고정한다 (스크롤해도 머리글이 보이게)
    #| 반복  기능마다 한 줄
        #| 갈래     모듈이 바뀌었나 ? 띠 색을 바꾼다 : 그대로
    #| 단계  자동 필터를 건다 — 모듈·영역·상태로 걸러 볼 수 있게
    #| 출력  기능목록 시트
    ws = wb.create_sheet("기능목록")
    _head(ws, cols)
    ws.freeze_panes = "E2"

    band, last = False, None
    for r, item in enumerate(rows, 2):
        if item["모듈"] != last:
            band, last = not band, item["모듈"]
        for c, key in enumerate(cols, 1):
            cell = ws.cell(row=r, column=c, value=item.get(key, ""))
            cell.font = Font(name=FONT, size=9.5, color=INK,
                             bold=(key == "기능명" and not item["함수"]))
            cell.alignment = Alignment(
                vertical="center", wrap_text=key in ("기능명", "설명(흐름)",
                                                     "입력", "출력", "비고"),
                horizontal="center" if key in ("번호", "단계", "분기", "반복",
                                               "호출", "상태", "줄") else "left")
            cell.border = BORDER
            if band:
                cell.fill = PatternFill("solid", fgColor=BAND)
        ws.row_dimensions[r].height = 30

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
    _widths(ws, [5, 12, 14, 20, 34, 40, 26, 26, 6, 6, 6, 6, 10, 6, 30])
    return ws


def sheet_modules(wb, modules, n_rows):
    """모듈 — 무엇을 맡는가 + 기능 수(수식)."""
    #| 흐름  모듈마다 역할과 기능 수를 쓴다. 수는 COUNTIF 수식으로.
    #| 반복  모듈마다
        #| 단계     기능 수 · 분기 합계를 기능목록에서 세는 수식을 넣는다
    #| 출력  모듈 시트
    ws = wb.create_sheet("모듈")
    cols = ["모듈(파일)", "영역", "맡는 일", "기능 수", "분기 합계", "단계 합계"]
    _head(ws, cols)
    ws.freeze_panes = "A2"

    end = n_rows + 1
    for r, (name, (area, role)) in enumerate(modules.items(), 2):
        vals = [name, area, role,
                f'=COUNTIF(기능목록!$C$2:$C${end},$A{r})',
                f'=SUMIF(기능목록!$C$2:$C${end},$A{r},기능목록!$J$2:$J${end})',
                f'=SUMIF(기능목록!$C$2:$C${end},$A{r},기능목록!$I$2:$I${end})']
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=9.5, color=INK,
                             bold=(c == 1))
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 3),
                                       horizontal="center" if c >= 4 else "left")
            cell.border = BORDER
        ws.row_dimensions[r].height = 24

    tot = len(modules) + 2
    ws.cell(row=tot, column=3, value="합계").font = Font(name=FONT, size=10, bold=True)
    for c in (4, 5, 6):
        L = get_column_letter(c)
        cell = ws.cell(row=tot, column=c, value=f"=SUM({L}2:{L}{tot - 1})")
        cell.font = Font(name=FONT, size=10, bold=True, color=INK)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    _widths(ws, [16, 14, 46, 10, 12, 12])
    return ws


def sheet_overview(wb, n_rows, n_mod):
    """개요 — 숫자는 전부 수식."""
    #| 흐름  제목과 요약 숫자를 쓴다. 숫자는 기능목록을 세는 수식.
    #| 출력  개요 시트
    ws = wb.create_sheet("개요", 0)
    end = n_rows + 1

    ws["A1"] = "🎻 바이올린 연습 도우미 — 기능 목록"
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color=INK)
    ws["A2"] = ("소스의 `#|` 흐름 주석에서 자동으로 뽑았습니다. "
                "`python3 features.py 기능리스트.xlsx` 로 다시 만들 수 있습니다.")
    ws["A2"].font = Font(name=FONT, size=9.5, color="5A6273")
    ws["A3"] = ("숫자는 손으로 적지 않고 수식입니다 — 기능목록이 바뀌면 여기도 같이 바뀝니다.")
    ws["A3"].font = Font(name=FONT, size=9.5, color="5A6273")

    items = [
        ("전체 기능 수", f"=COUNTA(기능목록!$A$2:$A${end})"),
        ("모듈 수", f"=COUNTA(모듈!$A$2:$A${n_mod + 1})"),
        ("완료", f'=COUNTIF(기능목록!$M$2:$M${end},"완료")'),
        ("부분 구현", f'=COUNTIF(기능목록!$M$2:$M${end},"부분 구현")'),
        ("분기(테스트해야 할 경우의 수)", f"=SUM(기능목록!$J$2:$J${end})"),
        ("단계 합계", f"=SUM(기능목록!$I$2:$I${end})"),
        ("다음 단계 항목", "=COUNTA(다음단계!$A$2:$A$100)"),
    ]
    for r, (label, formula) in enumerate(items, 5):
        a = ws.cell(row=r, column=1, value=label)
        a.font = Font(name=FONT, size=10, color=INK)
        a.border = BORDER
        b = ws.cell(row=r, column=2, value=formula)
        b.font = Font(name=FONT, size=12, bold=True, color=INK)
        b.alignment = Alignment(horizontal="center")
        b.border = BORDER
        b.fill = PatternFill("solid", fgColor=BAND)
        ws.row_dimensions[r].height = 22

    ws["A14"] = "시트 안내"
    ws["A14"].font = Font(name=FONT, size=11, bold=True, color=INK)
    guide = [("기능목록", "기능 한 줄씩. 모듈명·함수명·입력·출력·분기 수 포함"),
             ("모듈", "파일이 무엇을 맡는가 + 기능 수 (수식)"),
             ("다음단계", "아직 만들지 않은 화면·기능")]
    for r, (name, desc) in enumerate(guide, 15):
        ws.cell(row=r, column=1, value=name).font = Font(name=FONT, size=9.5,
                                                        bold=True, color=INK)
        ws.cell(row=r, column=2, value=desc).font = Font(name=FONT, size=9.5,
                                                         color=INK)
    _widths(ws, [34, 62])
    return ws


def sheet_backlog(wb, backlog):
    """다음단계 — 아직 안 만든 것."""
    #| 흐름  아직 안 만든 화면·기능을 우선순위와 함께 적는다
    #| 출력  다음단계 시트
    ws = wb.create_sheet("다음단계")
    cols = ["항목", "건드릴 모듈", "우선순위", "내용"]
    _head(ws, cols)
    ws.freeze_panes = "A2"
    for r, row in enumerate(backlog, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=9.5, color=INK, bold=(c == 1))
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 4),
                                       horizontal="center" if c == 3 else "left")
            cell.border = BORDER
        ws.row_dimensions[r].height = 34
    _widths(ws, [24, 24, 10, 70])
    return ws


def write(path, rows, modules, backlog):
    """엑셀 한 권을 씁니다."""
    #| 흐름  빈 책을 만들어 시트 네 장을 채우고 저장한다
    #| 입력  저장 경로 · 기능 목록 · 모듈 표 · 다음 단계
    #| 호출  sheet_features → 기능목록
    #| 호출  sheet_modules → 모듈
    #| 호출  sheet_backlog → 다음단계
    #| 호출  sheet_overview → 개요 (수식이 앞 시트들을 가리킵니다)
    #| 출력  .xlsx 파일
    from features import COLS
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_features(wb, rows, COLS)
    sheet_modules(wb, modules, len(rows))
    sheet_backlog(wb, backlog)
    sheet_overview(wb, len(rows), len(modules))
    wb.active = 0
    wb.save(path)
